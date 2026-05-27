import openpyxl
import json
from datetime import datetime

wb = openpyxl.load_workbook('CONTROLE BOLETOS OPERAÇÃO FIDC - BANCOS - 2025.xlsx', data_only=True)

sheet_map = {
    'DAYCOVAL': 'DAYCOVAL',
    'SAFRA': 'SAFRA',
    'MULTIPLIKE': 'MULTIPLIKE',
    'C6': 'C6',
    'CREDVALE': 'CREDVALE',
    'ALL SEC': 'ALL SEC',
    'AUDAX': 'AUDAX',
    'KANASTRA': 'KANASTRA',
    'YAALEH': 'YAALEH',
    'ONE 7': 'ONE 7',
    'JUMP': 'JUMP',
    'OPHIR': 'OPHIR',
    'KREDITON': 'KREDITON',
    'ASIA': 'ASIA',
    'SICOOB': 'SICOOB',
    'RED ASSET': 'RED ASSET',
    'MAIN3': 'MAIN3',
    'MANCHESTER': 'MANCHESTER',
    'ASA': 'ASA',
}

artico_name = next((s for s in wb.sheetnames if 'RTICO' in s), None)

results = {}

def process_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    header = rows[0]
    header_clean = [str(h).strip().upper() if h else '' for h in header]

    val_idx = next((i for i, h in enumerate(header_clean) if 'VALOR' in h and 'R$' in h), None)
    if val_idx is None:
        val_idx = next((i for i, h in enumerate(header_clean) if h.startswith('VALOR')), None)

    desagio_idx = next((i for i, h in enumerate(header_clean) if 'DESAGIO' in h), None)
    data_idx = next((i for i, h in enumerate(header_clean) if h == 'DATA'), None)
    venc_idx = next((i for i, h in enumerate(header_clean) if 'VENCIMENTO' in h), None)

    total_valor = 0.0
    total_desagio = 0.0
    datas = []
    vencimentos = []
    count = 0
    prazos = []

    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue

        valor = None
        desagio = None

        try:
            if val_idx is not None and row[val_idx] is not None:
                v = float(row[val_idx])
                if v > 0:
                    valor = v
        except:
            pass

        try:
            if desagio_idx is not None and row[desagio_idx] is not None:
                d = float(row[desagio_idx])
                # Filtrar deságio inválido: deve ser positivo e menor que o valor do boleto
                if d > 0 and valor is not None and d < valor:
                    desagio = d
        except:
            pass

        if valor is not None:
            total_valor += valor
            count += 1
            if desagio is not None:
                total_desagio += desagio

        try:
            if data_idx is not None and row[data_idx] is not None:
                d = row[data_idx]
                if isinstance(d, datetime):
                    datas.append(d)
        except:
            pass

        try:
            if venc_idx is not None and row[venc_idx] is not None:
                d = row[venc_idx]
                if isinstance(d, datetime):
                    vencimentos.append(d)
        except:
            pass

        try:
            di = row[data_idx] if data_idx is not None else None
            vi = row[venc_idx] if venc_idx is not None else None
            if isinstance(di, datetime) and isinstance(vi, datetime):
                prazo = (vi - di).days
                if 0 < prazo < 1000:
                    prazos.append(prazo)
        except:
            pass

    if total_valor == 0:
        return None

    data_inicio = min(datas).strftime('%d/%m/%Y') if datas else 'N/A'
    data_fim = max(vencimentos).strftime('%d/%m/%Y') if vencimentos else 'N/A'
    prazo_medio = int(sum(prazos) / len(prazos)) if prazos else 0

    return {
        'valor': round(total_valor, 2),
        'desagio': round(total_desagio, 2),
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'volumetria': count,
        'prazo_medio': prazo_medio,
    }

for sheet_key, display_name in sheet_map.items():
    if sheet_key in wb.sheetnames:
        result = process_sheet(wb[sheet_key])
        if result:
            results[display_name] = result

if artico_name:
    result = process_sheet(wb[artico_name])
    if result:
        results['ÁRTICO'] = result

# Corrigir chave ÁRTICO com encoding quebrado (openpyxl lê como caractere de substituição)
broken_keys = [k for k in results if 'RTICO' in k and k != 'ÁRTICO']
for bk in broken_keys:
    results['ÁRTICO'] = results.pop(bk)

# ─── GRAFENO: arquivo separado ────────────────────────────────────────────────
wb_g = openpyxl.load_workbook('CONTROLE BOLETOS  FIDC - GRAFENO - 2024.xlsx', data_only=True)
ws_g = wb_g['BOLETOS FIDC']
rows_g = list(ws_g.iter_rows(values_only=True))
hc_g = [str(h).strip().upper() if h else '' for h in rows_g[0]]
val_gi  = next((i for i,h in enumerate(hc_g) if 'VALOR' in h and 'R$' in h), None)
dat_gi  = next((i for i,h in enumerate(hc_g) if h == 'DATA'), None)
venc_gi = next((i for i,h in enumerate(hc_g) if 'VENCIMENTO' in h), None)
cedido_gi = next((i for i,h in enumerate(hc_g) if 'CEDIDO' in h), None)

g_valor = g_vol = 0; g_datas = []; g_vencs = []; g_prazos = []
for row in rows_g[1:]:
    if not any(v is not None for v in row): continue
    if cedido_gi is not None and not str(row[cedido_gi]).strip().upper().startswith('S'): continue
    try: v = float(row[val_gi]) if val_gi is not None and row[val_gi] is not None else 0.0
    except: v = 0.0
    if v <= 0: continue
    g_valor += v; g_vol += 1
    d = row[dat_gi]; vd = row[venc_gi]
    if isinstance(d, datetime): g_datas.append(d)
    if isinstance(vd, datetime): g_vencs.append(vd)
    if isinstance(d, datetime) and isinstance(vd, datetime):
        p = (vd - d).days
        if 0 < p < 1000: g_prazos.append(p)

if g_valor > 0:
    results['GRAFENO'] = {
        'valor': round(g_valor, 2), 'desagio': 0.0,
        'data_inicio': min(g_datas).strftime('%d/%m/%Y') if g_datas else 'N/A',
        'data_fim':    max(g_vencs).strftime('%d/%m/%Y') if g_vencs else 'N/A',
        'volumetria':  g_vol,
        'prazo_medio': round(sum(g_prazos)/len(g_prazos)) if g_prazos else 0,
    }

total_geral = sum(v['valor'] for v in results.values())
total_desagio_geral = sum(v['desagio'] for v in results.values())
total_boletos = sum(v['volumetria'] for v in results.values())

for name in results:
    r = results[name]
    r['representatividade'] = round(r['valor'] / total_geral * 100, 2) if total_geral > 0 else 0
    r['desagio_pct'] = round(r['desagio'] / r['valor'] * 100, 2) if r['valor'] > 0 else 0

results_sorted = dict(sorted(results.items(), key=lambda x: x[1]['valor'], reverse=True))

print("Dados coletados:")
for k, v in results_sorted.items():
    print(f"  {k}: R$ {v['valor']:,.0f} | Deságio: R$ {v['desagio']:,.0f} ({v['desagio_pct']:.2f}%) | {v['volumetria']} boletos | {v['representatividade']:.1f}% | Prazo médio: {v['prazo_medio']}d")
print(f"\nTOTAL: R$ {total_geral:,.0f} | Deságio: R$ {total_desagio_geral:,.0f} | {total_boletos} boletos")

data_export = {
    'fidics': results_sorted,
    'total': round(total_geral, 2),
    'total_desagio': round(total_desagio_geral, 2),
    'total_boletos': total_boletos,
    'gerado_em': datetime.now().strftime('%d/%m/%Y %H:%M')
}

with open('fidic_data.json', 'w', encoding='utf-8') as f:
    json.dump(data_export, f, ensure_ascii=False, indent=2)

print("\nDados salvos em fidic_data.json")
