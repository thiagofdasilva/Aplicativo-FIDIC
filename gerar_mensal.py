"""
Gera dados mensais por FIDC com status de pagamento do CSV de faturas recebidas.
Chave de cruzamento: str(NOTA_FISCAL).lstrip('0') + str(PARCELA).zfill(3)
Arquivo: FINFATURASRECEBIDASPORDATADERECEBIMENTOCTRECEBERACS.csv
Lógica: presente no arquivo = Pago; ausente = Não pago
"""
import openpyxl, json, csv, io
from datetime import datetime
from collections import defaultdict

# ─── CSV FATURAS RECEBIDAS ────────────────────────────────────────────────────
with open('FINFATURASRECEBIDASPORDATADERECEBIMENTOCTRECEBERACS.csv', 'r', encoding='utf-8', errors='replace') as f:
    content = f.read()

reader = csv.DictReader(io.StringIO(content))
fn = reader.fieldnames
col_nf  = fn[7]   # Número da Nota Fiscal
col_par = fn[15]  # Numero Parcela - ACS

# Conjunto de chaves pagas (presença = Pago, ausência = Não pago)
csv_lookup = set()
for r in reader:
    nf  = str(r[col_nf]).lstrip('0')
    par = str(r[col_par]).zfill(3)
    csv_lookup.add(nf + par)

print(f'CSV carregado: {len(csv_lookup):,} titulos pagos')

# ─── EXCEL ────────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(
    'CONTROLE BOLETOS OPERAÇÃO FIDC - BANCOS - 2025.xlsx', data_only=True
)

SHEET_NAMES = {
    'DAYCOVAL':'DAYCOVAL','SAFRA':'SAFRA','MULTIPLIKE':'MULTIPLIKE','C6':'C6',
    'CREDVALE':'CREDVALE','ALL SEC':'ALL SEC','AUDAX':'AUDAX','KANASTRA':'KANASTRA',
    'YAALEH':'YAALEH','ONE 7':'ONE 7','JUMP':'JUMP','OPHIR':'OPHIR',
    'KREDITON':'KREDITON','ASIA':'ASIA','SICOOB':'SICOOB','RED ASSET':'RED ASSET',
    'MAIN3':'MAIN3','MANCHESTER':'MANCHESTER','ASA':'ASA',
}
ARTICO = next((s for s in wb.sheetnames if 'RTICO' in s), None)

def process(ws, display_name):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}
    header = rows[0]
    hc = [str(h).strip().upper() if h else '' for h in header]

    val_idx  = next((i for i,h in enumerate(hc) if 'VALOR' in h and 'R$' in h), None)
    nf_idx   = next((i for i,h in enumerate(hc) if 'NOTA' in h or h == 'NF'), None)
    par_idx  = next((i for i,h in enumerate(hc) if 'PARCELA' in h), None)
    data_idx = next((i for i,h in enumerate(hc) if h == 'DATA'), None)
    venc_idx = next((i for i,h in enumerate(hc) if 'VENCIMENTO' in h), None)
    desagio_idx = next((i for i,h in enumerate(hc) if 'DESAGIO' in h), None)

    if val_idx is None or data_idx is None:
        print(f'  AVISO: colunas não encontradas em {display_name}')
        return {}

    monthly = defaultdict(lambda: {
        'valor': 0.0, 'desagio': 0.0, 'volumetria': 0, 'prazos': [],
        'v_pago': 0.0, 'v_aberto': 0.0, 'v_sem_info': 0.0,
        'n_pago': 0,   'n_aberto': 0,  'n_sem_info': 0,
    })
    total_check = 0.0
    matched = no_match = 0

    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        d = row[data_idx]
        if not isinstance(d, datetime):
            continue
        try:
            v = float(row[val_idx]) if row[val_idx] is not None else 0.0
        except (TypeError, ValueError):
            v = 0.0
        if v <= 0:
            continue

        # Chave de cruzamento: NF + PARCELA.zfill(3)
        try:
            nf_raw = str(row[nf_idx]).lstrip('0') if nf_idx is not None and row[nf_idx] is not None else ''
            par_raw = str(int(float(str(row[par_idx])))).zfill(3) if par_idx is not None and row[par_idx] is not None else '001'
            key = nf_raw + par_raw
        except:
            key = ''

        is_pago = key in csv_lookup

        ym = d.strftime('%Y-%m')
        m  = monthly[ym]
        m['valor']      += v
        m['volumetria'] += 1
        total_check     += v

        # Deságio
        if desagio_idx is not None and row[desagio_idx] is not None:
            try:
                des = float(row[desagio_idx])
                if 0 < des < v:
                    m['desagio'] += des
            except (TypeError, ValueError):
                pass

        # Prazo
        if venc_idx is not None and row[venc_idx] is not None:
            vd = row[venc_idx]
            if isinstance(vd, datetime):
                p = (vd - d).days
                if 0 < p < 1000:
                    m['prazos'].append(p)

        # Status
        if is_pago:
            m['v_pago']  += v
            m['n_pago']  += 1
            matched += 1
        else:
            m['v_aberto']  += v
            m['n_aberto']  += 1
            no_match += 1

    # Finalizar meses
    result = {}
    for ym, data in sorted(monthly.items()):
        ps = data.pop('prazos')
        result[ym] = {
            'valor':      round(data['valor'], 2),
            'desagio':    round(data['desagio'], 2),
            'volumetria': data['volumetria'],
            'prazo_medio': round(sum(ps) / len(ps)) if ps else 0,
            'v_pago':     round(data['v_pago'], 2),
            'v_aberto':   round(data['v_aberto'], 2),
            'v_sem_info': round(data['v_sem_info'], 2),
            'n_pago':     data['n_pago'],
            'n_aberto':   data['n_aberto'],
            'n_sem_info': data['n_sem_info'],
        }

    total_emb = sum(m['valor'] for m in result.values())
    pct = round(matched/(matched+no_match)*100,1) if (matched+no_match) else 0
    print(f'  {display_name:<15} total={total_emb:>15,.2f}  match={pct}%  '
          f'pago={sum(m["n_pago"] for m in result.values())}  '
          f'aberto={sum(m["n_aberto"] for m in result.values())}  '
          f'sem_info={sum(m["n_sem_info"] for m in result.values())}')
    return result


print('\nProcessando FIDICs...')
all_monthly = {}
for sk, dn in SHEET_NAMES.items():
    if sk in wb.sheetnames:
        m = process(wb[sk], dn)
        if m:
            all_monthly[dn] = m

if ARTICO:
    m = process(wb[ARTICO], 'ARTICO')
    if m:
        all_monthly['ARTICO'] = m

# ─── GRAFENO: arquivo separado, status via coluna SITUAÇÃO ────────────────────
wb_g = openpyxl.load_workbook('CONTROLE BOLETOS  FIDC - GRAFENO - 2024.xlsx', data_only=True)
ws_g = wb_g['BOLETOS FIDC']
rows_g = list(ws_g.iter_rows(values_only=True))
hc_g = [str(h).strip().upper() if h else '' for h in rows_g[0]]
val_gi   = next((i for i,h in enumerate(hc_g) if 'VALOR' in h and 'R$' in h), None)
dat_gi   = next((i for i,h in enumerate(hc_g) if h == 'DATA'), None)
venc_gi  = next((i for i,h in enumerate(hc_g) if 'VENCIMENTO' in h), None)
cedido_gi= next((i for i,h in enumerate(hc_g) if 'CEDIDO' in h), None)
sit_gi   = next((i for i,h in enumerate(hc_g) if 'SITUA' in h), None)
des_gi   = next((i for i,h in enumerate(hc_g) if 'DES' in h and 'GIO' in h), None)

monthly_g = defaultdict(lambda: {
    'valor':0.0,'desagio':0.0,'volumetria':0,'prazos':[],
    'v_pago':0.0,'v_aberto':0.0,'v_sem_info':0.0,
    'n_pago':0,'n_aberto':0,'n_sem_info':0,
})
n_pg_g = n_ab_g = 0
for row in rows_g[1:]:
    if not any(v is not None for v in row): continue
    if cedido_gi is not None and not str(row[cedido_gi]).strip().upper().startswith('S'): continue
    d = row[dat_gi]
    if not isinstance(d, datetime): continue
    try: v = float(row[val_gi]) if val_gi is not None and row[val_gi] is not None else 0.0
    except: v = 0.0
    if v <= 0: continue

    try:
        des_g = float(row[des_gi]) if des_gi is not None and isinstance(row[des_gi], (int, float)) else 0.0
        if des_g >= v: des_g = 0.0
    except: des_g = 0.0

    sit = str(row[sit_gi]).strip().upper() if sit_gi is not None and row[sit_gi] else ''
    is_pago = 'BAIXADO' in sit

    ym = d.strftime('%Y-%m')
    m = monthly_g[ym]
    m['valor'] += v; m['desagio'] += des_g; m['volumetria'] += 1
    if venc_gi is not None and isinstance(row[venc_gi], datetime):
        p = (row[venc_gi] - d).days
        if 0 < p < 1000: m['prazos'].append(p)
    if is_pago:
        m['v_pago']  += v; m['n_pago']  += 1; n_pg_g += 1
    else:
        m['v_aberto'] += v; m['n_aberto'] += 1; n_ab_g += 1

result_g = {}
for ym, data in sorted(monthly_g.items()):
    ps = data.pop('prazos')
    result_g[ym] = {
        'valor': round(data['valor'],2), 'desagio': round(data['desagio'],2), 'volumetria': data['volumetria'],
        'prazo_medio': round(sum(ps)/len(ps)) if ps else 0,
        'v_pago': round(data['v_pago'],2), 'v_aberto': round(data['v_aberto'],2), 'v_sem_info': 0.0,
        'n_pago': data['n_pago'], 'n_aberto': data['n_aberto'], 'n_sem_info': 0,
    }
if result_g:
    all_monthly['GRAFENO'] = result_g
    total_g = sum(m['valor'] for m in result_g.values())
    pct_g = round(n_pg_g/(n_pg_g+n_ab_g)*100,1) if (n_pg_g+n_ab_g) else 0
    print(f'  {"GRAFENO":<15} total={total_g:>15,.2f}  match={pct_g}%  pago={n_pg_g}  aberto={n_ab_g}  sem_info=0')

all_months = sorted({ym for fund in all_monthly.values() for ym in fund})

out = {'monthly': all_monthly, 'all_months': all_months}
with open('fidic_monthly_v2.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, separators=(',', ':'))

sz = len(json.dumps(out, ensure_ascii=False, separators=(',', ':')))
print(f'\nSalvo: fidic_monthly_v2.json  ({sz/1024:.1f} KB)')
