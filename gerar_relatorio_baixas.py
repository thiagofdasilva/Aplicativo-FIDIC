"""
Relatório de baixas Jan/Fev/Mar 2026
Lógica: JOIN entre planilha FIDC + ACS Faturas via chave NF.lstrip('0') + PARCELA.zfill(3)
Data de referência: VENCIMENTO (no mês alvo)
Status de pagamento: ACS Faturas CSV
"""
import openpyxl, csv, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from collections import defaultdict

# ─── CONFIG ───────────────────────────────────────────────────────────────────
MESES_ALVO  = {(2026, 1), (2026, 2), (2026, 3)}
NOMES_MESES = {1: 'Janeiro/2026', 2: 'Fevereiro/2026', 3: 'Março/2026'}
HOJE = datetime.today()

# ─── STEP 1: montar lookup ACS Faturas ────────────────────────────────────────
with open('ACS Faturas.csv', 'r', encoding='utf-8', errors='replace') as f:
    content = f.read()
reader = csv.DictReader(io.StringIO(content))
fn = reader.fieldnames
col_nf, col_par = fn[1], fn[3]
col_val_csv, col_rst, col_deb, col_sts = fn[6], fn[7], fn[8], fn[17]

col_venc_csv = fn[5]   # Data de vencimento no CSV (para calcular data pagamento)

csv_lkp = {}   # chave -> {status, valor_csv, restante, dias_deb, data_pag_est}
for r in reader:
    nf  = str(r[col_nf]).lstrip('0')
    par = str(r[col_par]).zfill(3)
    key = nf + par
    try:    vc  = float(r[col_val_csv])
    except: vc  = 0.0
    try:    rst = float(r[col_rst])
    except: rst = 0.0
    try:    deb = int(float(r[col_deb]))
    except: deb = 0
    st = r[col_sts].strip()
    if st.lower().startswith('n') and 'pago' in st.lower():
        st = 'Não pago'
    # Estimar data de pagamento: vencimento + dias_em_debito (só para Status=Pago)
    data_pag_est = None
    if st == 'Pago':
        try:
            venc_csv = datetime.strptime(r[col_venc_csv].strip(), '%d/%m/%Y')
            data_pag_est = venc_csv + timedelta(days=deb)
        except:
            data_pag_est = None
    csv_lkp[key] = {'status': st, 'valor_csv': vc, 'restante': rst,
                    'dias_deb': deb, 'data_pag_est': data_pag_est}

print(f'CSV: {len(csv_lkp):,} registros carregados')

# ─── STEP 2: varrer planilha FIDC e fazer join ────────────────────────────────
wb_src = openpyxl.load_workbook(
    'CONTROLE BOLETOS OPERAÇÃO FIDC - BANCOS - 2025.xlsx', data_only=True
)
SHEET_NAMES = {
    'DAYCOVAL':'DAYCOVAL','SAFRA':'SAFRA','MULTIPLIKE':'MULTIPLIKE','C6':'C6',
    'CREDVALE':'CREDVALE','ALL SEC':'ALL SEC','AUDAX':'AUDAX','KANASTRA':'KANASTRA',
    'YAALEH':'YAALEH','ONE 7':'ONE 7','JUMP':'JUMP','OPHIR':'OPHIR',
    'KREDITON':'KREDITON','ASIA':'ASIA','SICOOB':'SICOOB','RED ASSET':'RED ASSET',
    'MAIN3':'MAIN3','MANCHESTER':'MANCHESTER','ASA':'ASA',
}
ARTICO = next((s for s in wb_src.sheetnames if 'RTICO' in s), None)
if ARTICO: SHEET_NAMES[ARTICO] = 'ÁRTICO'

registros = []

for sk, display in SHEET_NAMES.items():
    if sk not in wb_src.sheetnames:
        continue
    ws = wb_src[sk]
    rows = list(ws.iter_rows(values_only=True))
    hc = [str(h).strip().upper() if h else '' for h in rows[0]]

    nf_idx   = next((i for i,h in enumerate(hc) if 'NOTA' in h or h == 'NF'), None)
    par_idx  = next((i for i,h in enumerate(hc) if 'PARCELA' in h), None)
    val_idx  = next((i for i,h in enumerate(hc) if 'VALOR' in h and 'R$' in h), None)
    dat_idx  = next((i for i,h in enumerate(hc) if h == 'DATA'), None)
    venc_idx = next((i for i,h in enumerate(hc) if 'VENCIMENTO' in h), None)
    cli_idx  = next((i for i,h in enumerate(hc) if 'CLIENTE' in h), None)
    cnpj_idx = next((i for i,h in enumerate(hc) if 'CNPJ' in h), None)
    des_idx  = next((i for i,h in enumerate(hc) if 'DESAGIO' in h), None)
    sit_idx  = next((i for i,h in enumerate(hc) if 'SITUA' in h), None)
    drec_idx = next((i for i,h in enumerate(hc) if 'DATA' in h and ('RECOMPRA' in h or 'BAIXA' in h)), None)

    if val_idx is None or venc_idx is None:
        continue

    n_reg = 0
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue

        # VENCIMENTO no mês alvo
        venc = row[venc_idx]
        if not isinstance(venc, datetime):
            continue
        if (venc.year, venc.month) not in MESES_ALVO:
            continue

        try:
            v = float(row[val_idx]) if row[val_idx] is not None else 0.0
        except:
            v = 0.0
        if v <= 0:
            continue

        # Chave JOIN
        try:
            nf_raw  = str(row[nf_idx]).lstrip('0') if nf_idx is not None and row[nf_idx] is not None else ''
            par_raw = str(int(float(str(row[par_idx])))).zfill(3) if par_idx is not None and row[par_idx] is not None else '001'
            key = nf_raw + par_raw
        except:
            key = ''

        # Join com ACS
        info = csv_lkp.get(key, {})
        status   = info.get('status', 'Sem registro')
        restante = info.get('restante', 0.0)
        dias_deb = info.get('dias_deb', 0)

        # Data de pagamento: 1ª opção DATA RECOMPRA da planilha; 2ª Venc+Dias (CSV)
        data_rec = row[drec_idx] if drec_idx is not None and row[drec_idx] is not None else None
        if isinstance(data_rec, datetime):
            data_pagamento = data_rec
            fonte_pag = 'DATA RECOMPRA'
        elif info.get('data_pag_est') and status == 'Pago':
            data_pagamento = info['data_pag_est']
            fonte_pag = 'Estimado (Venc.+Dias)'
        else:
            data_pagamento = None
            fonte_pag = '—'

        # Deságio
        try:
            des = float(row[des_idx]) if des_idx is not None and row[des_idx] is not None else 0.0
            if des >= v: des = 0.0
        except:
            des = 0.0

        # Data cessão
        data_cess = row[dat_idx] if dat_idx is not None and isinstance(row[dat_idx], datetime) else None

        # NF como inteiro
        try:
            nf_int = int(nf_raw) if nf_raw.isdigit() else nf_raw
        except:
            nf_int = nf_raw

        registros.append({
            'fidc':         display,
            'mes':          venc.month,
            'mes_nome':     NOMES_MESES[venc.month],
            'data_cessao':  data_cess,
            'vencimento':   venc,
            'data_pag':     data_pagamento,
            'fonte_pag':    fonte_pag,
            'cliente':      str(row[cli_idx]) if cli_idx is not None and row[cli_idx] is not None else '',
            'cnpj':         str(row[cnpj_idx]) if cnpj_idx is not None and row[cnpj_idx] is not None else '',
            'nf':           nf_int,
            'parcela':      int(par_raw),
            'valor':        round(v, 2),
            'desagio':      round(des, 2),
            'val_liq':      round(v - des, 2),
            'restante':     round(restante, 2),
            'dias_deb':     dias_deb,
            'status':       status,
            'sit_interna':  str(row[sit_idx]).strip() if sit_idx is not None and row[sit_idx] is not None else '',
        })
        n_reg += 1

    print(f'  {display:<15}: {n_reg:>5} registros com vencimento em jan-mar/2026')

# Ordenar: FIDC → mês → vencimento
registros.sort(key=lambda r: (r['fidc'], r['mes'], r['vencimento']))
print(f'\nTotal geral: {len(registros):,} registros')

# ─── STEP 3: estatísticas ─────────────────────────────────────────────────────
# por fidc+mês: valor, desagio, n, n_pago, n_npago
Stats = lambda: {'n':0,'valor':0,'desagio':0,'val_liq':0,'n_pago':0,'n_npago':0,'n_sem':0}
stats = defaultdict(lambda: defaultdict(Stats))
for r in registros:
    s = stats[r['fidc']][r['mes']]
    s['n']       += 1
    s['valor']   += r['valor']
    s['desagio'] += r['desagio']
    s['val_liq'] += r['val_liq']
    if r['status']   == 'Pago':        s['n_pago']  += 1
    elif r['status'] == 'Não pago':    s['n_npago'] += 1
    else:                              s['n_sem']   += 1

# ─── ESTILOS ──────────────────────────────────────────────────────────────────
def fill(h): return PatternFill('solid', fgColor=h)
def fnt(bold=False, color='E2E8F0', size=9, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name='Calibri')
def brd():
    s = Side(style='thin', color='2E3354')
    return Border(top=s, bottom=s, left=s, right=s)
def aln(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

C = dict(
    bg='0F1117', bg2='1A1D26', bg3='21253A',
    acc='4F6EF7', grn='10B981', yel='F59E0B',
    red='EF4444', pur='7C3AED', txt='E2E8F0',
    txt2='94A3B8', txt3='64748B', wht='FFFFFF',
)
MES_ACC = {1:'3B82F6', 2:'0D9488', 3:'EA580C'}
MES_BG  = {1:'0D1F3C', 2:'0D2626', 3:'2D1A0E'}

# ─── WORKBOOK ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

def hdr_row(ws, row, cols, texts, bg, fg='FFFFFF', sz=9, bold=True, h=26):
    ws.row_dimensions[row].height = h
    for ci, t in enumerate(texts, cols[0]):
        c = ws.cell(row=row, column=ci, value=t)
        c.font = Font(bold=bold, color=fg, size=sz, name='Calibri')
        c.fill = fill(bg); c.border = brd(); c.alignment = aln('center')

def titulo_ws(ws, row, text, bg, ncols, sz=14, h=40):
    ws.row_dimensions[row].height = h
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws['A' + str(row)]
    c.value = '  ' + text
    c.font = Font(bold=True, color=C['wht'], size=sz, name='Calibri')
    c.fill = fill(bg); c.alignment = aln('left')

def sep_mes(ws, row, mes, ncols):
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    c = ws.cell(row=row, column=1, value=f'  ▶  {NOMES_MESES[mes].upper()}')
    c.font = Font(bold=True, color=C['wht'], size=10, name='Calibri')
    c.fill = fill(MES_ACC[mes]); c.alignment = aln('left')

# ════════════════════════════════════════════════════════════════════════════════
# ABA 1 — RESUMO CONSOLIDADO
# ════════════════════════════════════════════════════════════════════════════════
ws_r = wb.active
ws_r.title = 'Resumo Consolidado'
ws_r.sheet_view.showGridLines = False
ws_r.sheet_properties.tabColor = C['acc']

for ci, w in enumerate([3,22,16,16,16,16,14,14,10,10,10], 1):
    ws_r.column_dimensions[get_column_letter(ci)].width = w

titulo_ws(ws_r, 2, 'BAIXAS — JAN / FEV / MAR 2026  |  RESUMO CONSOLIDADO', C['acc'], 11)
ws_r.row_dimensions[3].height = 18
ws_r.merge_cells('A3:K3')
ws_r['A3'] = f'  Gerado em {HOJE.strftime("%d/%m/%Y %H:%M")}  •  {len(registros):,} registros  •  Fonte: VENCIMENTO (planilha FIDC) + Status (ACS Faturas)'
ws_r['A3'].font = Font(color=C['txt2'], size=8, italic=True, name='Calibri')
ws_r['A3'].fill = fill(C['bg3']); ws_r['A3'].alignment = aln('left')

hdr_row(ws_r, 5, [1], ['FIDC','Jan - Valor','Jan - Deságio','Fev - Valor','Fev - Deságio',
                        'Mar - Valor','Mar - Deságio','TOTAL Valor','TOTAL Deságio','Pagos','Não Pagos'], C['bg3'])

fidcs = sorted(stats.keys())
tv = td = tvj = tdj = tvf = tdf = tvm = tdm = 0
for ri, fidc in enumerate(fidcs, 6):
    ws_r.row_dimensions[ri].height = 20
    bg = fill('15192B') if ri%2==0 else fill(C['bg2'])
    j=stats[fidc][1]; f=stats[fidc][2]; m=stats[fidc][3]
    vj=j['valor']; dj=j['desagio']; vf=f['valor']; df=f['desagio']; vm=m['valor']; dm=m['desagio']
    vt=vj+vf+vm; dt=dj+df+dm
    np=j['n_pago']+f['n_pago']+m['n_pago']
    nn=j['n_npago']+f['n_npago']+m['n_npago']+j['n_sem']+f['n_sem']+m['n_sem']
    vals=[fidc,vj,dj,vf,df,vm,dm,vt,dt,np,nn]
    for ci, val in enumerate(vals, 1):
        c = ws_r.cell(row=ri, column=ci, value=round(val,2) if isinstance(val,float) else val)
        c.fill = bg; c.border = brd()
        if ci==1:
            c.font=Font(bold=True,color=C['txt'],size=9,name='Calibri'); c.alignment=aln('left')
        elif ci in(2,4,6,8):
            c.font=fnt(color='93C5FD'); c.number_format='R$ #,##0.00'; c.alignment=aln('right')
        elif ci in(3,5,7,9):
            c.font=fnt(color='FBBF24'); c.number_format='R$ #,##0.00'; c.alignment=aln('right')
        elif ci==10:
            c.font=fnt(color=C['grn']); c.alignment=aln('center')
        else:
            c.font=fnt(color=C['red']); c.alignment=aln('center')
    tv+=vt; td+=dt; tvj+=vj; tdj+=dj; tvf+=vf; tdf+=df; tvm+=vm; tdm+=dm

tr=len(fidcs)+6
ws_r.row_dimensions[tr].height=24
tp=sum(stats[f][m]['n_pago'] for f in fidcs for m in(1,2,3))
tn=sum(stats[f][m]['n_npago']+stats[f][m]['n_sem'] for f in fidcs for m in(1,2,3))
for ci,val in enumerate(['TOTAL GERAL',tvj,tdj,tvf,tdf,tvm,tdm,tv,td,tp,tn],1):
    c=ws_r.cell(row=tr,column=ci,value=round(val,2) if isinstance(val,float) else val)
    c.fill=fill(C['acc']); c.font=Font(bold=True,color=C['wht'],size=9,name='Calibri')
    c.border=brd(); c.alignment=aln('center' if ci>1 else 'left')
    if ci in(2,4,6,8): c.number_format='R$ #,##0.00'
    if ci in(3,5,7,9): c.number_format='R$ #,##0.00'

# ════════════════════════════════════════════════════════════════════════════════
# ABA 2 — DESÁGIO p/ CONTABILIDADE (por fundo × mês)
# ════════════════════════════════════════════════════════════════════════════════
ws_cont = wb.create_sheet('Desagio p. Contabilidade')
ws_cont.sheet_view.showGridLines = False
ws_cont.sheet_properties.tabColor = C['yel']

for ci,w in enumerate([3,22,20,20,20,20],1):
    ws_cont.column_dimensions[get_column_letter(ci)].width = w

titulo_ws(ws_cont, 2, 'DESÁGIO POR FUNDO E MÊS — CLASSIFICAÇÃO CONTA DE JUROS', C['yel'], 6, sz=13)
hdr_row(ws_cont, 4, [1], ['FIDC','Jan/2026','Fev/2026','Mar/2026','TOTAL','% Deságio'], C['bg3'])

for ri,fidc in enumerate(fidcs,5):
    ws_cont.row_dimensions[ri].height=20
    bg=fill('15192B') if ri%2==0 else fill(C['bg2'])
    dj=stats[fidc][1]['desagio']; df=stats[fidc][2]['desagio']; dm=stats[fidc][3]['desagio']
    dt=dj+df+dm
    vt=stats[fidc][1]['valor']+stats[fidc][2]['valor']+stats[fidc][3]['valor']
    pct=round(dt/vt*100,2) if vt else 0
    for ci,val in enumerate([fidc,dj,df,dm,dt,pct],1):
        c=ws_cont.cell(row=ri,column=ci,value=round(val,2) if isinstance(val,float) else val)
        c.fill=bg; c.border=brd()
        if ci==1:
            c.font=Font(bold=True,color=C['txt'],size=9,name='Calibri'); c.alignment=aln('left')
        elif ci==6:
            c.font=fnt(color='A78BFA'); c.number_format='0.00"%"'; c.alignment=aln('right')
        else:
            c.font=fnt(color='FBBF24'); c.number_format='R$ #,##0.00'; c.alignment=aln('right')

tr_c=len(fidcs)+5
ws_cont.row_dimensions[tr_c].height=24
dj2=sum(stats[f][1]['desagio'] for f in fidcs)
df2=sum(stats[f][2]['desagio'] for f in fidcs)
dm2=sum(stats[f][3]['desagio'] for f in fidcs)
dt2=dj2+df2+dm2; vt2=tv
pct2=round(dt2/vt2*100,2) if vt2 else 0
for ci,val in enumerate(['TOTAL DESÁGIO',dj2,df2,dm2,dt2,pct2],1):
    c=ws_cont.cell(row=tr_c,column=ci,value=round(val,2) if isinstance(val,float) else val)
    c.fill=fill(C['yel']); c.font=Font(bold=True,color='0F1117',size=9,name='Calibri')
    c.border=brd(); c.alignment=aln('center' if ci>1 else 'left')
    if 2<=ci<=5: c.number_format='R$ #,##0.00'
    if ci==6: c.number_format='0.00"%"'

# ════════════════════════════════════════════════════════════════════════════════
# ABA 3 — DETALHAMENTO COMPLETO
# ════════════════════════════════════════════════════════════════════════════════
ws_det = wb.create_sheet('Detalhamento Completo')
ws_det.sheet_view.showGridLines = False
ws_det.sheet_properties.tabColor = C['grn']
ws_det.freeze_panes = 'A3'

COL_W = [12,38,18,8,7,12,12,14,10,15,14,14,12,18,12]
COL_H = ['FIDC','Cliente','CNPJ','NF','Parcela','Data Cessão','Vencimento',
         'Data Pagamento','Dias p/ Venc.','Valor (R$)','Deságio (R$)','Valor Líq. (R$)',
         'Status ACS','Fonte Data Pag.','Situação']
for i,w in enumerate(COL_W,1):
    ws_det.column_dimensions[get_column_letter(i)].width = w

titulo_ws(ws_det, 1, f'BAIXAS JAN/FEV/MAR 2026 — DETALHAMENTO COMPLETO  ({len(registros):,} registros)', C['grn'], len(COL_H))
hdr_row(ws_det, 2, [1], COL_H, C['bg3'], sz=8, h=26)

cur_mes = None; dr = 3
for r in registros:
    if r['mes'] != cur_mes:
        cur_mes = r['mes']
        sep_mes(ws_det, dr, cur_mes, len(COL_H))
        dr += 1
    ws_det.row_dimensions[dr].height = 16
    bg_c = fill('14181F') if dr%2==0 else fill(C['bg2'])
    sc   = C['grn'] if r['status']=='Pago' else (C['red'] if r['status']=='Não pago' else C['txt3'])
    fc   = '7C3AED' if r['fonte_pag']=='DATA RECOMPRA' else ('94A3B8' if r['fonte_pag']=='Estimado (Venc.+Dias)' else C['txt3'])
    dias_ref = (r['vencimento'] - HOJE).days if r['vencimento'] else None
    #         1-FIDC  2-Cliente  3-CNPJ  4-NF  5-Parcela  6-Cessão  7-Venc  8-DataPag  9-Dias  10-Valor  11-Deságio  12-Líq  13-Status  14-FontePag  15-Sit
    vals=[r['fidc'],r['cliente'],r['cnpj'],r['nf'],r['parcela'],
          r['data_cessao'],r['vencimento'],r['data_pag'],dias_ref,
          r['valor'],r['desagio'] if r['desagio'] else None,
          r['val_liq'],r['status'],r['fonte_pag'],r['sit_interna']]
    for ci,val in enumerate(vals,1):
        c=ws_det.cell(row=dr,column=ci,value=val)
        c.fill=bg_c; c.border=brd()
        if ci==1:   c.font=Font(bold=True,color=C['txt'],size=8,name='Calibri'); c.alignment=aln('left')
        elif ci==2: c.font=fnt(color=C['txt'],size=8); c.alignment=aln('left')
        elif ci==3: c.font=fnt(color=C['txt2'],size=8); c.alignment=aln('left')
        elif ci in(4,5): c.font=fnt(color=C['txt2'],size=8); c.alignment=aln('center')
        elif ci in(6,7): c.font=fnt(color=C['txt2'],size=8); c.alignment=aln('center'); c.number_format='DD/MM/YYYY'
        elif ci==8:
            pag_color = C['grn'] if r['fonte_pag']=='DATA RECOMPRA' else '94A3B8'
            c.font=Font(bold=r['fonte_pag']=='DATA RECOMPRA',color=pag_color,size=8,name='Calibri')
            c.alignment=aln('center'); c.number_format='DD/MM/YYYY'
        elif ci==9:
            col=C['grn'] if dias_ref and dias_ref>=0 else (C['yel'] if dias_ref and dias_ref>=-30 else C['red'])
            c.font=Font(bold=False,color=col,size=8,name='Calibri'); c.alignment=aln('center')
        elif ci==10: c.font=fnt(color='93C5FD',size=8); c.number_format='#,##0.00'; c.alignment=aln('right')
        elif ci==11: c.font=fnt(color='FBBF24',size=8); c.number_format='#,##0.00'; c.alignment=aln('right')
        elif ci==12: c.font=Font(bold=True,color=C['txt'],size=8,name='Calibri'); c.number_format='#,##0.00'; c.alignment=aln('right')
        elif ci==13: c.font=Font(bold=True,color=sc,size=8,name='Calibri'); c.alignment=aln('center')
        elif ci==14: c.font=fnt(color=fc,size=8); c.alignment=aln('center')
        elif ci==15: c.font=fnt(color=C['txt2'],size=8); c.alignment=aln('center')
    dr += 1

# Linha total
ws_det.row_dimensions[dr].height=22
tv_d=sum(r['valor'] for r in registros)
td_d=sum(r['desagio'] for r in registros)
tl_d=sum(r['val_liq'] for r in registros)
n_pg =sum(1 for r in registros if r['status']=='Pago')
n_np =sum(1 for r in registros if r['status']=='Não pago')
n_si =len(registros)-n_pg-n_np
n_rec=sum(1 for r in registros if r['fonte_pag']=='DATA RECOMPRA')
n_est=sum(1 for r in registros if r['fonte_pag']=='Estimado (Venc.+Dias)')
tot=[f'TOTAL — {len(registros):,} registros',
     f'Pagos: {n_pg} | N.pagos: {n_np} | S/reg: {n_si}',
     '','','','','','',f'Rec.:{n_rec} Est.:{n_est}',tv_d,td_d,tl_d,'','','']
for ci,val in enumerate(tot,1):
    c=ws_det.cell(row=dr,column=ci,value=val)
    c.fill=fill(C['bg3']); c.font=Font(bold=True,color=C['wht'],size=8,name='Calibri')
    c.border=brd(); c.alignment=aln('right' if ci>=10 else 'left')
    if ci in(9,10,11): c.number_format='#,##0.00'

# ════════════════════════════════════════════════════════════════════════════════
# ABA POR FIDC
# ════════════════════════════════════════════════════════════════════════════════
for fidc in fidcs:
    regs = [r for r in registros if r['fidc']==fidc]
    if not regs: continue
    safe = fidc[:31]
    ws_f = wb.create_sheet(safe)
    ws_f.sheet_view.showGridLines = False
    ws_f.sheet_properties.tabColor = C['acc']
    ws_f.freeze_panes = 'A3'
    nc = len(COL_H)-1   # sem coluna FIDC
    HDRS_F = COL_H[1:]

    for i,w in enumerate(COL_W[1:],1):
        ws_f.column_dimensions[get_column_letter(i)].width=w

    titulo_ws(ws_f, 1, f'{fidc}  — Vencimentos Jan/Fev/Mar 2026  ({len(regs)} registros)', C['acc'], nc)
    hdr_row(ws_f, 2, [1], HDRS_F, C['bg3'], sz=8, h=26)

    # Sub-totais por mês
    for mes in (1,2,3):
        rm = [r for r in regs if r['mes']==mes]
        if not rm: continue
        sep_mes(ws_f, ws_f.max_row+1, mes, nc)
        for r in rm:
            ri_f = ws_f.max_row+1
            ws_f.row_dimensions[ri_f].height=16
            bg_c=fill('14181F') if ri_f%2==0 else fill(C['bg2'])
            sc=C['grn'] if r['status']=='Pago' else (C['red'] if r['status']=='Não pago' else C['txt3'])
            dias_ref=(r['vencimento']-HOJE).days if r['vencimento'] else None
            fc_=('7C3AED' if r['fonte_pag']=='DATA RECOMPRA' else
                 ('94A3B8' if r['fonte_pag']=='Estimado (Venc.+Dias)' else C['txt3']))
            #    1-Cli  2-CNPJ  3-NF  4-Parc  5-Cessão  6-Venc  7-DataPag  8-Dias  9-Valor  10-Deság  11-Líq  12-Status  13-Fonte  14-Sit
            vals_f=[r['cliente'],r['cnpj'],r['nf'],r['parcela'],
                    r['data_cessao'],r['vencimento'],r['data_pag'],dias_ref,
                    r['valor'],r['desagio'] if r['desagio'] else None,
                    r['val_liq'],r['status'],r['fonte_pag'],r['sit_interna']]
            for ci,val in enumerate(vals_f,1):
                c=ws_f.cell(row=ri_f,column=ci,value=val)
                c.fill=bg_c; c.border=brd()
                if ci==1: c.font=fnt(color=C['txt'],size=8); c.alignment=aln('left')
                elif ci==2: c.font=fnt(color=C['txt2'],size=8); c.alignment=aln('left')
                elif ci in(3,4): c.font=fnt(color=C['txt2'],size=8); c.alignment=aln('center')
                elif ci in(5,6): c.font=fnt(color=C['txt2'],size=8); c.alignment=aln('center'); c.number_format='DD/MM/YYYY'
                elif ci==7:
                    pag_c=C['grn'] if r['fonte_pag']=='DATA RECOMPRA' else '94A3B8'
                    c.font=Font(bold=r['fonte_pag']=='DATA RECOMPRA',color=pag_c,size=8,name='Calibri')
                    c.alignment=aln('center'); c.number_format='DD/MM/YYYY'
                elif ci==8:
                    col=C['grn'] if dias_ref and dias_ref>=0 else (C['yel'] if dias_ref and dias_ref>=-30 else C['red'])
                    c.font=Font(bold=False,color=col,size=8,name='Calibri'); c.alignment=aln('center')
                elif ci==9:  c.font=fnt(color='93C5FD',size=8); c.number_format='#,##0.00'; c.alignment=aln('right')
                elif ci==10: c.font=fnt(color='FBBF24',size=8); c.number_format='#,##0.00'; c.alignment=aln('right')
                elif ci==11: c.font=Font(bold=True,color=C['txt'],size=8,name='Calibri'); c.number_format='#,##0.00'; c.alignment=aln('right')
                elif ci==12: c.font=Font(bold=True,color=sc,size=8,name='Calibri'); c.alignment=aln('center')
                elif ci==13: c.font=fnt(color=fc_,size=8); c.alignment=aln('center')
                elif ci==14: c.font=fnt(color=C['txt2'],size=8); c.alignment=aln('center')

        # Sub-total por mês
        ri_st = ws_f.max_row+1
        ws_f.row_dimensions[ri_st].height=20
        vm_=sum(r['valor'] for r in rm)
        dm_=sum(r['desagio'] for r in rm)
        lm_=sum(r['val_liq'] for r in rm)
        np_=sum(1 for r in rm if r['status']=='Pago')
        nr_=sum(1 for r in rm if r['fonte_pag']=='DATA RECOMPRA')
        ne_=sum(1 for r in rm if r['fonte_pag']=='Estimado (Venc.+Dias)')
        for ci,val in enumerate([f'Sub-total {NOMES_MESES[mes]}: {len(rm)} registros',
                                  '','','','','',f'Pagos:{np_}/{len(rm)}',
                                  f'Rec:{nr_} Est:{ne_}',vm_,dm_,lm_,'','',''],1):
            c=ws_f.cell(row=ri_st,column=ci,value=round(val,2) if isinstance(val,float) else val)
            c.fill=fill(MES_BG[mes]); c.font=Font(bold=True,color=C['wht'],size=8,name='Calibri')
            c.border=brd(); c.alignment=aln('right' if ci>=9 else 'left')
            if ci in(9,10,11): c.number_format='#,##0.00'

    # Total geral da aba
    ri_tot=ws_f.max_row+1; ws_f.row_dimensions[ri_tot].height=22
    tvf_=sum(r['valor'] for r in regs)
    tdf_=sum(r['desagio'] for r in regs)
    tlf_=sum(r['val_liq'] for r in regs)
    npf_=sum(1 for r in regs if r['status']=='Pago')
    nrf_=sum(1 for r in regs if r['fonte_pag']=='DATA RECOMPRA')
    nef_=sum(1 for r in regs if r['fonte_pag']=='Estimado (Venc.+Dias)')
    for ci,val in enumerate([f'TOTAL {fidc}: {len(regs)} registros',
                              '','','','','',f'Pagos:{npf_}/{len(regs)}',
                              f'Rec:{nrf_} Est:{nef_}',tvf_,tdf_,tlf_,'','',''],1):
        c=ws_f.cell(row=ri_tot,column=ci,value=round(val,2) if isinstance(val,float) else val)
        c.fill=fill(C['acc']); c.font=Font(bold=True,color=C['wht'],size=9,name='Calibri')
        c.border=brd(); c.alignment=aln('right' if ci>=9 else 'left')
        if ci in(9,10,11): c.number_format='#,##0.00'

# ─── ORDENAR ABAS ─────────────────────────────────────────────────────────────
wb.move_sheet('Resumo Consolidado',    offset=-len(wb.sheetnames)+1)
wb.move_sheet('Desagio p. Contabilidade', offset=-len(wb.sheetnames)+2)
wb.move_sheet('Detalhamento Completo', offset=-len(wb.sheetnames)+3)

# ─── SALVAR ───────────────────────────────────────────────────────────────────
ts = HOJE.strftime('%Y%m%d_%H%M')
filename = f'Baixas_Jan_Fev_Mar_2026_{ts}.xlsx'
wb.save(filename)

print(f'\nArquivo: {filename}')
print(f'  Registros: {len(registros):,}')
print(f'  Valor total:   R$ {tv:,.2f}')
print(f'  Desagio total: R$ {td:,.2f}')
print(f'  Pagos: {n_pg} | Nao pagos: {n_np} | Sem registro: {n_si}')
print(f'  Abas: Resumo + Desagio + Detalhamento + {len(fidcs)} FIDCs')

# ─── STEP 5: Gerar baixas_raw.js para o dashboard ─────────────────────────────
import json as _json
VENC_MIN = datetime(2025, 1, 1)

cli_dict_js = {}
cli_arr_js  = []
records_js  = []
months_set  = set()

for sk, display in SHEET_NAMES.items():
    if sk not in wb_src.sheetnames:
        continue
    ws2 = wb_src[sk]
    rows2 = list(ws2.iter_rows(values_only=True))
    hc2 = [str(h).strip().upper() if h else '' for h in rows2[0]]

    nf2   = next((i for i,h in enumerate(hc2) if 'NOTA' in h or h == 'NF'), None)
    par2  = next((i for i,h in enumerate(hc2) if 'PARCELA' in h), None)
    val2  = next((i for i,h in enumerate(hc2) if 'VALOR' in h and 'R$' in h), None)
    dat2  = next((i for i,h in enumerate(hc2) if h == 'DATA'), None)
    vnc2  = next((i for i,h in enumerate(hc2) if 'VENCIMENTO' in h), None)
    cli2  = next((i for i,h in enumerate(hc2) if 'CLIENTE' in h), None)
    des2  = next((i for i,h in enumerate(hc2) if 'DESAGIO' in h), None)
    drec2 = next((i for i,h in enumerate(hc2) if 'DATA' in h and ('RECOMPRA' in h or 'BAIXA' in h)), None)

    if val2 is None or vnc2 is None:
        continue

    for row in rows2[1:]:
        if not any(v is not None for v in row):
            continue
        venc = row[vnc2]
        if not isinstance(venc, datetime) or venc < VENC_MIN:
            continue
        try:
            v = float(row[val2]) if row[val2] is not None else 0.0
        except:
            v = 0.0
        if v <= 0:
            continue

        try:
            nf_r = str(row[nf2]).lstrip('0') if nf2 is not None and row[nf2] is not None else ''
            pa_r = str(int(float(str(row[par2])))).zfill(3) if par2 is not None and row[par2] is not None else '001'
            key2 = nf_r + pa_r
        except:
            nf_r, pa_r, key2 = '', '001', ''

        info2 = csv_lkp.get(key2, {})
        st2   = info2.get('status', 'Sem registro')
        if st2 == 'Pago':
            st_enc = 'P'
        elif st2 and 'pago' in st2.lower() and st2.strip()[0].lower() == 'n':
            st_enc = 'N'
        else:
            st_enc = 'S'

        # Data pagamento
        dr2 = row[drec2] if drec2 is not None and row[drec2] is not None else None
        if isinstance(dr2, datetime):
            dp_str = dr2.strftime('%Y-%m-%d')
        elif info2.get('data_pag_est') and st2 == 'Pago':
            dp_str = info2['data_pag_est'].strftime('%Y-%m-%d')
        else:
            dp_str = None

        try:
            des_v = float(row[des2]) if des2 is not None and row[des2] is not None else 0.0
            if des_v >= v: des_v = 0.0
        except:
            des_v = 0.0

        cess_str = row[dat2].strftime('%Y-%m-%d') if dat2 is not None and isinstance(row[dat2], datetime) else None
        nf_js = int(nf_r) if nf_r.isdigit() else nf_r

        cli_name = str(row[cli2]) if cli2 is not None and row[cli2] is not None else ''
        if cli_name not in cli_dict_js:
            cli_dict_js[cli_name] = len(cli_arr_js)
            cli_arr_js.append(cli_name)

        mes_str = venc.strftime('%Y-%m')
        months_set.add(mes_str)
        records_js.append([
            display, cli_dict_js[cli_name], mes_str, venc.day,
            cess_str, nf_js, int(pa_r),
            round(v, 2), round(des_v, 2), st_enc, dp_str,
        ])

js_obj = {
    'h': ['FIDC','Cli','MesVenc','DiaVenc','Cessao','NF','Parcela','Valor','Desagio','Status','DataPag'],
    'd': records_js,
    'c': cli_arr_js,
    'm': sorted(months_set),
}
js_content = 'var BAIXAS_RAW=' + _json.dumps(js_obj, ensure_ascii=False, separators=(',', ':')) + ';'
with open('baixas_raw.js', 'w', encoding='utf-8') as f:
    f.write(js_content)
print(f'\nbaixas_raw.js: {len(records_js):,} registros, {len(months_set)} meses')
