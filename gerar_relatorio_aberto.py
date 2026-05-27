"""
Gera relatório Excel de títulos em aberto cruzando planilha FIDC x Faturas Recebidas ACS
"""
import openpyxl, csv, io
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, date

# ─── CSV FATURAS RECEBIDAS (Pago = presente; ausente = Não pago) ──────────────
with open('FINFATURASRECEBIDASPORDATADERECEBIMENTOCTRECEBERACS.csv', 'r', encoding='utf-8', errors='replace') as f:
    content = f.read()

reader = csv.DictReader(io.StringIO(content))
fn = reader.fieldnames
col_nf  = fn[7]   # Número da Nota Fiscal
col_par = fn[15]  # Numero Parcela - ACS

# Conjunto de chaves pagas — ausência = aberto
csv_lookup = set()
for r in reader:
    nf  = str(r[col_nf]).lstrip('0')
    par = str(r[col_par]).zfill(3)
    csv_lookup.add(nf + par)

# ─── EXCEL FIDIC ──────────────────────────────────────────────────────────────
wb_src = openpyxl.load_workbook(
    'CONTROLE BOLETOS OPERAÇÃO FIDC - BANCOS - 2025.xlsx', data_only=True
)

SHEET_NAMES = {
    'DAYCOVAL':'DAYCOVAL','SAFRA':'SAFRA','MULTIPLIKE':'MULTIPLIKE','C6':'C6',
    'CREDVALE':'CREDVALE','ALL SEC':'ALL SEC','AUDAX':'AUDAX','KANASTRA':'KANASTRA',
    'YAALEH':'YAALEH','ONE 7':'ONE 7','JUMP':'JUMP','OPHIR':'OPHIR',
    'KREDITON':'KREDITON','ASIA':'ASIA','SICOOB':'SICOOB','RED ASSET':'RED ASSET',
    'MAIN3':'MAIN3','MANCHESTER':'MANCHESTER','ASA':'ASA',
}
ARTICO = next((s for s in wb_src.sheetnames if 'RTICO' in s), None)
if ARTICO:
    SHEET_NAMES[ARTICO] = 'ÁRTICO'

# ─── COLETA TÍTULOS EM ABERTO ─────────────────────────────────────────────────
registros = []
hoje = datetime.today()

for sk, display in SHEET_NAMES.items():
    if sk not in wb_src.sheetnames:
        continue
    ws = wb_src[sk]
    rows = list(ws.iter_rows(values_only=True))
    hc = [str(h).strip().upper() if h else '' for h in rows[0]]

    nf_idx   = next((i for i,h in enumerate(hc) if 'NOTA' in h or h == 'NF'), None)
    par_idx  = next((i for i,h in enumerate(hc) if 'PARCELA' in h), None)
    val_idx  = next((i for i,h in enumerate(hc) if 'VALOR' in h and 'R$' in h), None)
    dat_idx  = next((i for i,h in enumerate(hc) if h == 'DATA'), None)
    venc_idx = next((i for i,h in enumerate(hc) if 'VENCIMENTO' in h), None)
    cli_idx  = next((i for i,h in enumerate(hc) if 'CLIENTE' in h), None)
    cnpj_idx = next((i for i,h in enumerate(hc) if 'CNPJ' in h), None)
    des_idx  = next((i for i,h in enumerate(hc) if 'DESAGIO' in h), None)

    if val_idx is None or nf_idx is None:
        continue

    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        try:
            v = float(row[val_idx]) if row[val_idx] is not None else 0.0
        except:
            v = 0.0
        if v <= 0:
            continue

        try:
            nf_raw = str(row[nf_idx]).lstrip('0') if row[nf_idx] is not None else ''
            par_raw = str(int(float(str(row[par_idx])))).zfill(3) if par_idx is not None and row[par_idx] is not None else '001'
            key = nf_raw + par_raw
        except:
            key = ''

        # Título pago = encontrado no CSV de recebidos → pular
        if key in csv_lookup:
            continue
        status = 'Não pago'

        # datas
        data_cess = row[dat_idx] if dat_idx is not None else None
        vencimento = row[venc_idx] if venc_idx is not None else None

        dias_venc = None
        if isinstance(vencimento, datetime):
            dias_venc = (hoje - vencimento).days
        elif isinstance(vencimento, date):
            dias_venc = (hoje.date() - vencimento).days

        try:
            desagio = float(row[des_idx]) if des_idx is not None and row[des_idx] is not None else 0.0
            if desagio >= v: desagio = 0.0
        except:
            desagio = 0.0

        registros.append({
            'fidc':      display,
            'cliente':   str(row[cli_idx]) if cli_idx is not None and row[cli_idx] is not None else '',
            'cnpj':      str(row[cnpj_idx]) if cnpj_idx is not None and row[cnpj_idx] is not None else '',
            'nf':        int(nf_raw) if nf_raw.isdigit() else nf_raw,
            'parcela':   int(par_raw) if par_raw.isdigit() else 1,
            'valor':     v,
            'desagio':   desagio,
            'valor_liq': round(v - desagio, 2),
            'restante':  v,
            'data_cess': data_cess,
            'vencimento': vencimento,
            'dias_venc': dias_venc,
            'status':    status,
        })

# ─── GRAFENO: arquivo separado, em aberto = SITUAÇÃO NORMAL ──────────────────
wb_g = openpyxl.load_workbook('CONTROLE BOLETOS  FIDC - GRAFENO - 2024.xlsx', data_only=True)
ws_g = wb_g['BOLETOS FIDC']
rows_g = list(ws_g.iter_rows(values_only=True))
hc_g = [str(h).strip().upper() if h else '' for h in rows_g[0]]
nf_gi    = next((i for i,h in enumerate(hc_g) if 'NOTA' in h or h == 'NF'), None)
par_gi   = next((i for i,h in enumerate(hc_g) if 'PARCELA' in h), None)
val_gi   = next((i for i,h in enumerate(hc_g) if 'VALOR' in h and 'R$' in h), None)
dat_gi   = next((i for i,h in enumerate(hc_g) if h == 'DATA'), None)
venc_gi  = next((i for i,h in enumerate(hc_g) if 'VENCIMENTO' in h), None)
cli_gi   = next((i for i,h in enumerate(hc_g) if 'CLIENTE' in h), None)
cnpj_gi  = next((i for i,h in enumerate(hc_g) if 'CNPJ' in h), None)
sit_gi   = next((i for i,h in enumerate(hc_g) if 'SITUA' in h), None)
cedido_gi= next((i for i,h in enumerate(hc_g) if 'CEDIDO' in h), None)
des_gi   = next((i for i,h in enumerate(hc_g) if 'DES' in h and 'GIO' in h), None)

for row in rows_g[1:]:
    if not any(v is not None for v in row): continue
    if cedido_gi is not None and not str(row[cedido_gi]).strip().upper().startswith('S'): continue
    sit = str(row[sit_gi]).strip().upper() if sit_gi is not None and row[sit_gi] else ''
    if 'BAIXADO' in sit: continue  # já pago, não é aberto
    try: v = float(row[val_gi]) if val_gi is not None and row[val_gi] is not None else 0.0
    except: v = 0.0
    if v <= 0: continue

    try:
        des_gv = float(row[des_gi]) if des_gi is not None and isinstance(row[des_gi], (int, float)) else 0.0
        if des_gv >= v: des_gv = 0.0
    except: des_gv = 0.0

    try:
        nf_raw = str(row[nf_gi]).lstrip('0') if nf_gi is not None and row[nf_gi] is not None else ''
        par_raw = str(int(float(str(row[par_gi])))).zfill(3) if par_gi is not None and row[par_gi] is not None else '001'
    except:
        nf_raw, par_raw = '', '001'

    data_cess = row[dat_gi] if dat_gi is not None and isinstance(row[dat_gi], datetime) else None
    vencimento = row[venc_gi] if venc_gi is not None else None
    dias_venc = None
    if isinstance(vencimento, datetime):
        dias_venc = (hoje - vencimento).days
    elif isinstance(vencimento, date):
        dias_venc = (hoje.date() - vencimento).days

    registros.append({
        'fidc':       'GRAFENO',
        'cliente':    str(row[cli_gi]) if cli_gi is not None and row[cli_gi] is not None else '',
        'cnpj':       str(row[cnpj_gi]) if cnpj_gi is not None and row[cnpj_gi] is not None else '',
        'nf':         int(nf_raw) if nf_raw.isdigit() else nf_raw,
        'parcela':    int(par_raw) if par_raw.isdigit() else 1,
        'valor':      v,
        'desagio':    round(des_gv, 2),
        'valor_liq':  round(v - des_gv, 2),
        'restante':   v,
        'data_cess':  data_cess,
        'vencimento': vencimento,
        'dias_venc':  dias_venc,
        'status':     'Não pago',
    })

# Ordenar: FIDC, depois vencimento mais antigo primeiro
registros.sort(key=lambda r: (
    r['fidc'],
    r['dias_venc'] if r['dias_venc'] is not None else -9999
), reverse=False)
registros.sort(key=lambda r: r['fidc'])

print(f'Títulos em aberto encontrados: {len(registros)}')

# ─── CRIAR EXCEL ──────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── CORES E ESTILOS ──
C_DARK     = '0F1117'
C_BG       = '1A1D26'
C_HEADER   = '21253A'
C_ACCENT   = '4F6EF7'
C_GREEN    = '10B981'
C_YELLOW   = 'F59E0B'
C_RED      = 'EF4444'
C_ORANGE   = 'F97316'
C_TEXT     = 'E2E8F0'
C_TEXT2    = '94A3B8'
C_WHITE    = 'FFFFFF'
C_PARCIAL  = '7C3AED'

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def font(bold=False, color=C_WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name='Calibri')

def border_thin(sides='all'):
    s = Side(style='thin', color='2E3354')
    n = Side(style=None)
    t = s if sides in ('all','top') else n
    b = s if sides in ('all','bottom') else n
    l = s if sides in ('all','left') else n
    r = s if sides in ('all','right') else n
    return Border(top=t, bottom=b, left=l, right=r)

def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def dias_color(dias):
    if dias is None: return C_TEXT2
    if dias <= 0:    return C_GREEN      # ainda no prazo / vence hoje
    if dias <= 30:   return C_YELLOW
    if dias <= 90:   return C_ORANGE
    return C_RED

# ════════════════════════════════════════════════════════════════════════════════
# ABA 1 — RESUMO EXECUTIVO
# ════════════════════════════════════════════════════════════════════════════════
ws_res = wb.active
ws_res.title = 'Resumo'
ws_res.sheet_view.showGridLines = False
ws_res.sheet_properties.tabColor = C_ACCENT

# Larguras
for col, w in zip('ABCDEFGH', [4, 28, 20, 20, 18, 18, 15, 20]):
    ws_res.column_dimensions[get_column_letter(ord(col)-64)].width = w

# ── Cabeçalho geral ──
ws_res.row_dimensions[1].height = 10
ws_res.row_dimensions[2].height = 45
ws_res.merge_cells('B2:H2')
ws_res['B2'] = '  RELATÓRIO DE TÍTULOS EM ABERTO — FIDCs'
ws_res['B2'].font = Font(bold=True, color=C_WHITE, size=18, name='Calibri')
ws_res['B2'].fill = fill(C_ACCENT)
ws_res['B2'].alignment = align('left')

ws_res.row_dimensions[3].height = 22
ws_res.merge_cells('B3:H3')
ws_res['B3'] = f'  Gerado em {hoje.strftime("%d/%m/%Y %H:%M")}  •  {len(registros)} títulos em aberto'
ws_res['B3'].font = Font(color=C_TEXT2, size=10, name='Calibri', italic=True)
ws_res['B3'].fill = fill(C_HEADER)
ws_res['B3'].alignment = align('left')

# ── Totalizadores por FIDC ──
from collections import defaultdict
por_fidc = defaultdict(lambda: {'n':0,'valor':0,'restante':0,'venc_30':0,'venc_90':0,'venc_mais':0})
for r in registros:
    f = por_fidc[r['fidc']]
    f['n']        += 1
    f['valor']    += r['valor']
    f['restante'] += r['restante'] if r['restante'] else r['valor']
    d = r['dias_venc']
    if d is not None:
        if d > 90:   f['venc_mais'] += 1
        elif d > 30: f['venc_90']   += 1
        elif d > 0:  f['venc_30']   += 1

# Cabeçalho da tabela de resumo
ws_res.row_dimensions[5].height = 8
ws_res.row_dimensions[6].height = 32
header_res = ['FIDC', 'Boletos Abertos', 'Valor em Aberto', 'Venc. até 30d', 'Venc. 31–90d', 'Venc. >90d']
for ci, h in enumerate(header_res, 2):
    c = ws_res.cell(row=6, column=ci, value=h)
    c.font = Font(bold=True, color=C_WHITE, size=10, name='Calibri')
    c.fill = fill(C_HEADER)
    c.alignment = align('center')
    c.border = border_thin()

total_n = total_v = total_v30 = total_v90 = total_vm = 0
fidcs_order = sorted(por_fidc.keys())
for ri, fidc in enumerate(fidcs_order, 7):
    f = por_fidc[fidc]
    ws_res.row_dimensions[ri].height = 22
    row_fill = fill('15192B') if ri % 2 == 0 else fill(C_BG)

    vals = [fidc, f['n'], f['valor'], f['venc_30'], f['venc_90'], f['venc_mais']]
    for ci, val in enumerate(vals, 2):
        c = ws_res.cell(row=ri, column=ci, value=val)
        c.fill = row_fill
        c.border = border_thin()
        c.alignment = align('center' if ci > 2 else 'left')
        if ci == 2:
            c.font = Font(bold=True, color=C_TEXT, size=10, name='Calibri')
        elif ci == 3:
            c.font = font(color=C_TEXT, size=10)
        elif ci == 4:
            c.font = font(color='93C5FD', size=10)
            c.number_format = 'R$ #,##0.00'
        elif ci == 5:
            c.font = font(color=C_YELLOW if f['venc_30'] > 0 else C_TEXT2, size=10)
        elif ci == 6:
            c.font = font(color=C_ORANGE if f['venc_90'] > 0 else C_TEXT2, size=10)
        elif ci == 7:
            c.font = font(color=C_RED if f['venc_mais'] > 0 else C_TEXT2, size=10)

    total_n  += f['n'];  total_v  += f['valor']
    total_v30 += f['venc_30']; total_v90 += f['venc_90']; total_vm += f['venc_mais']

# Linha total
tr = len(fidcs_order) + 7
ws_res.row_dimensions[tr].height = 26
for ci, val in enumerate(['TOTAL GERAL', total_n, total_v, total_v30, total_v90, total_vm], 2):
    c = ws_res.cell(row=tr, column=ci, value=val)
    c.fill = fill(C_ACCENT)
    c.font = Font(bold=True, color=C_WHITE, size=10, name='Calibri')
    c.alignment = align('center' if ci > 2 else 'left')
    c.border = border_thin()
    if ci == 4: c.number_format = 'R$ #,##0.00'

# ════════════════════════════════════════════════════════════════════════════════
# ABA 2 — TÍTULOS DETALHADOS
# ════════════════════════════════════════════════════════════════════════════════
ws_det = wb.create_sheet('Títulos em Aberto')
ws_det.sheet_view.showGridLines = False
ws_det.sheet_properties.tabColor = C_RED
ws_det.freeze_panes = 'A3'

# Larguras das colunas
COL_WIDTHS = [14, 38, 20, 10, 10, 16, 16, 14, 14, 14, 16, 16]
COL_HEADERS = [
    'FIDC', 'Cliente', 'CNPJ', 'NF', 'Parcela',
    'Data Cessão', 'Vencimento', 'Dias Vencido',
    'Valor (R$)', 'Deságio (R$)', 'Valor Líq. (R$)', 'Status'
]
for i, w in enumerate(COL_WIDTHS, 1):
    ws_det.column_dimensions[get_column_letter(i)].width = w

# Linha título
ws_det.row_dimensions[1].height = 36
ws_det.merge_cells(f'A1:{get_column_letter(len(COL_HEADERS))}1')
ws_det['A1'] = '  TÍTULOS EM ABERTO — DETALHAMENTO COMPLETO'
ws_det['A1'].font = Font(bold=True, color=C_WHITE, size=14, name='Calibri')
ws_det['A1'].fill = fill(C_RED)
ws_det['A1'].alignment = align('left')

# Cabeçalho colunas
ws_det.row_dimensions[2].height = 30
for ci, h in enumerate(COL_HEADERS, 1):
    c = ws_det.cell(row=2, column=ci, value=h)
    c.font = Font(bold=True, color=C_WHITE, size=9, name='Calibri')
    c.fill = fill(C_HEADER)
    c.alignment = align('center')
    c.border = border_thin()

# Dados
for ri, r in enumerate(registros, 3):
    ws_det.row_dimensions[ri].height = 18
    row_fill = fill('15192B') if ri % 2 == 0 else fill(C_BG)
    dv = r['dias_venc']
    dc = dias_color(dv)
    st = r['status']
    st_color = C_RED if st == 'Não pago' else C_PARCIAL

    vals = [
        r['fidc'],
        r['cliente'],
        r['cnpj'],
        r['nf'],
        r['parcela'],
        r['data_cess'],
        r['vencimento'],
        dv,
        r['valor'],
        r['desagio'] if r['desagio'] else None,
        r['valor_liq'],
        st,
    ]
    for ci, val in enumerate(vals, 1):
        c = ws_det.cell(row=ri, column=ci, value=val)
        c.fill = row_fill
        c.border = border_thin()
        c.alignment = align('left' if ci <= 3 else 'center')

        if ci == 1:     # FIDC
            c.font = Font(bold=True, color=C_TEXT, size=9, name='Calibri')
        elif ci == 2:   # Cliente
            c.font = Font(color=C_TEXT, size=9, name='Calibri')
            c.alignment = align('left')
        elif ci == 3:   # CNPJ
            c.font = Font(color=C_TEXT2, size=9, name='Calibri')
        elif ci in (4,5):  # NF, Parcela
            c.font = Font(color=C_TEXT2, size=9, name='Calibri')
        elif ci in (6,7):  # Datas
            c.font = Font(color=C_TEXT2, size=9, name='Calibri')
            c.number_format = 'DD/MM/YYYY'
        elif ci == 8:   # Dias vencido
            c.font = Font(bold=dv is not None and dv > 0, color=dc, size=9, name='Calibri')
            c.alignment = align('center')
        elif ci == 9:   # Valor
            c.font = Font(color='93C5FD', size=9, name='Calibri')
            c.number_format = '#,##0.00'
        elif ci == 10:  # Deságio
            c.font = Font(color=C_YELLOW, size=9, name='Calibri')
            c.number_format = '#,##0.00'
        elif ci == 11:  # Valor líq.
            c.font = Font(bold=True, color=C_TEXT, size=9, name='Calibri')
            c.number_format = '#,##0.00'
        elif ci == 12:  # Status
            c.font = Font(bold=True, color=st_color, size=9, name='Calibri')

# Linha de totais detalhados
tr_det = len(registros) + 3
ws_det.row_dimensions[tr_det].height = 24
total_val  = sum(r['valor']    for r in registros)
total_des  = sum(r['desagio']  for r in registros)
total_liq  = sum(r['valor_liq']for r in registros)

tot_vals = ['TOTAL', f'{len(registros)} títulos', '', '', '', '', '', '',
            total_val, total_des if total_des else None, total_liq, '']
for ci, val in enumerate(tot_vals, 1):
    c = ws_det.cell(row=tr_det, column=ci, value=val)
    c.fill = fill(C_HEADER)
    c.font = Font(bold=True, color=C_WHITE, size=9, name='Calibri')
    c.alignment = align('center' if ci > 3 else 'left')
    c.border = border_thin()
    if ci == 9:  c.number_format = '#,##0.00'
    if ci == 10: c.number_format = '#,##0.00'
    if ci == 11: c.number_format = '#,##0.00'

# ════════════════════════════════════════════════════════════════════════════════
# ABAS POR FIDC (com pelo menos 1 título em aberto)
# ════════════════════════════════════════════════════════════════════════════════
for fidc in fidcs_order:
    regs_fidc = [r for r in registros if r['fidc'] == fidc]
    if not regs_fidc:
        continue

    safe_name = fidc[:31]  # Excel limita a 31 chars
    ws_f = wb.create_sheet(safe_name)
    ws_f.sheet_view.showGridLines = False
    ws_f.sheet_properties.tabColor = C_ACCENT
    ws_f.freeze_panes = 'A3'

    for i, w in enumerate(COL_WIDTHS[1:], 1):  # sem coluna FIDC
        ws_f.column_dimensions[get_column_letter(i)].width = w

    # Título
    headers_f = COL_HEADERS[1:]  # sem FIDC
    ncols = len(headers_f)
    ws_f.row_dimensions[1].height = 36
    ws_f.merge_cells(f'A1:{get_column_letter(ncols)}1')
    ws_f['A1'] = f'  {fidc} — Títulos em Aberto ({len(regs_fidc)} boletos)'
    ws_f['A1'].font = Font(bold=True, color=C_WHITE, size=13, name='Calibri')
    ws_f['A1'].fill = fill(C_ACCENT)
    ws_f['A1'].alignment = align('left')

    ws_f.row_dimensions[2].height = 28
    for ci, h in enumerate(headers_f, 1):
        c = ws_f.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, color=C_WHITE, size=9, name='Calibri')
        c.fill = fill(C_HEADER)
        c.alignment = align('center')
        c.border = border_thin()

    for ri, r in enumerate(regs_fidc, 3):
        ws_f.row_dimensions[ri].height = 18
        row_fill = fill('15192B') if ri % 2 == 0 else fill(C_BG)
        dv = r['dias_venc']
        dc = dias_color(dv)
        st_color = C_RED if r['status'] == 'Não pago' else C_PARCIAL

        vals_f = [
            r['cliente'], r['cnpj'], r['nf'], r['parcela'],
            r['data_cess'], r['vencimento'], dv,
            r['valor'], r['desagio'] if r['desagio'] else None,
            r['valor_liq'], r['status'],
        ]
        for ci, val in enumerate(vals_f, 1):
            c = ws_f.cell(row=ri, column=ci, value=val)
            c.fill = row_fill
            c.border = border_thin()
            if ci == 1:
                c.font = Font(color=C_TEXT, size=9, name='Calibri')
                c.alignment = align('left')
            elif ci == 2:
                c.font = Font(color=C_TEXT2, size=9, name='Calibri')
                c.alignment = align('left')
            elif ci in (3,4):
                c.font = Font(color=C_TEXT2, size=9, name='Calibri')
                c.alignment = align('center')
            elif ci in (5,6):
                c.font = Font(color=C_TEXT2, size=9, name='Calibri')
                c.alignment = align('center')
                c.number_format = 'DD/MM/YYYY'
            elif ci == 7:
                c.font = Font(bold=dv is not None and dv > 0, color=dc, size=9, name='Calibri')
                c.alignment = align('center')
            elif ci == 8:
                c.font = Font(color='93C5FD', size=9, name='Calibri')
                c.number_format = '#,##0.00'
                c.alignment = align('right')
            elif ci == 9:
                c.font = Font(color=C_YELLOW, size=9, name='Calibri')
                c.number_format = '#,##0.00'
                c.alignment = align('right')
            elif ci == 10:
                c.font = Font(bold=True, color=C_TEXT, size=9, name='Calibri')
                c.number_format = '#,##0.00'
                c.alignment = align('right')
            elif ci == 11:
                c.font = Font(bold=True, color=st_color, size=9, name='Calibri')
                c.alignment = align('center')

    # Total da aba
    tr_f = len(regs_fidc) + 3
    ws_f.row_dimensions[tr_f].height = 24
    tv = sum(r['valor']    for r in regs_fidc)
    td = sum(r['desagio']  for r in regs_fidc)
    tl = sum(r['valor_liq']for r in regs_fidc)
    for ci, val in enumerate(['TOTAL', '', '', '', '', '', '',
                               tv, td if td else None, tl, ''], 1):
        c = ws_f.cell(row=tr_f, column=ci, value=val)
        c.fill = fill(C_HEADER)
        c.font = Font(bold=True, color=C_WHITE, size=9, name='Calibri')
        c.alignment = align('right' if ci >= 8 else 'left')
        c.border = border_thin()
        if ci in (8,9,10): c.number_format = '#,##0.00'

# Mover Resumo para primeira posição
wb.move_sheet('Resumo', offset=-len(wb.sheetnames)+1)

# Salvar
ts = hoje.strftime('%Y%m%d_%H%M')
filename = f'Relatorio_Abertos_{ts}.xlsx'
wb.save(filename)
print(f'\nRelatorio salvo: {filename}')
print(f'  Titulos em aberto: {len(registros)}')
print(f'  Valor total:       R$ {total_val:,.2f}')
print(f'  FIDCs com abertos: {len(fidcs_order)}')

# ─── Gerar abertos_raw.js para o dashboard ────────────────────────────────────
import json as _json
ab_records = []
for r in registros:
    venc_str = r['vencimento'].strftime('%Y-%m-%d') if isinstance(r['vencimento'], datetime) else (
               r['vencimento'].strftime('%Y-%m-%d') if isinstance(r['vencimento'], date) else None)
    cess_str = r['data_cess'].strftime('%Y-%m-%d') if isinstance(r['data_cess'], datetime) else None
    ab_records.append([
        r['fidc'], r['cliente'], r['cnpj'],
        r['nf'], r['parcela'],
        cess_str, venc_str,
        r['dias_venc'],
        round(r['valor'], 2),
        round(r['desagio'], 2),
        round(r['valor_liq'], 2),
        'Nao pago',
    ])
ab_obj = {
    'h': ['FIDC','Cliente','CNPJ','NF','Parcela','Data Cessao','Vencimento',
          'Dias Vencido','Valor','Desagio','Valor Liq','Status'],
    'd': ab_records,
}
js_ab = 'var ABERTOS_RAW=' + _json.dumps(ab_obj, ensure_ascii=False, separators=(',',':')) + ';'
with open('abertos_raw.js', 'w', encoding='utf-8') as f:
    f.write(js_ab)
print(f'  abertos_raw.js: {len(ab_records):,} registros')
